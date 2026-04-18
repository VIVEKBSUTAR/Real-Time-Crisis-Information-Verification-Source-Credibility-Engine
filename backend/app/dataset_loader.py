"""
Dataset Loader for Bharat Fake News Dataset
Loads Excel file into memory for fast similarity matching
"""
import json
import os
from typing import List, Dict, Tuple
import logging

logger = logging.getLogger(__name__)


def normalize_label(label_value) -> int:
    """Convert any label format to 0 (false) or 1 (true)"""
    if label_value is None:
        return 0
    
    # Handle various types
    label_str = str(label_value).lower().strip()
    
    # String representations
    if label_str in ['1', 'true', 'yes', 'verified', 'fact', 'correct', 'real']:
        return 1
    elif label_str in ['0', 'false', 'no', 'fake', 'debunked', 'incorrect', 'incorrect', 'misinformation']:
        return 0
    
    # Numeric
    try:
        return 1 if int(float(label_str)) == 1 else 0
    except:
        pass
    
    # Default to 0 if unclear
    return 0


class DatasetLoader:
    """Load and manage fake news dataset"""
    
    def __init__(self, excel_path: str = None):
        self.excel_path = excel_path or "/Volumes/V_Mac_SSD/Hackathon/Breaking Enigma/bharatfakenewskosh (3).xlsx"
        self.claims = []
        self.loaded = False
    
    def load(self) -> bool:
        """Load dataset from Excel file"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        
        if not os.path.exists(self.excel_path):
            logger.error(f"Dataset file not found: {self.excel_path}")
            return False
        
        try:
            print(f"📂 Loading dataset from: {self.excel_path}")
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Load claims
            claim_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, row))
                
                # Extract key fields
                claim_id = row_dict.get('id', f'CLAIM_{row_idx}')
                claim_text = row_dict.get('Eng_Trans_Statement') or row_dict.get('Statement', '')
                raw_label = row_dict.get('Label', 0)
                label = normalize_label(raw_label)  # ← FIX: Normalize all labels
                source = row_dict.get('Fact_Check_Source', 'Unknown')
                
                if claim_text.strip():
                    self.claims.append({
                        'id': claim_id,
                        'text': claim_text.strip(),
                        'label': label,  # Now always 0 or 1
                        'source': source,
                        'author': row_dict.get('Author_Name', ''),
                        'category': row_dict.get('News_Category', '')
                    })
                    claim_count += 1
                    
                    # Progress every 5000 rows
                    if claim_count % 5000 == 0:
                        print(f"   ✓ Loaded {claim_count:,} claims...")
            
            self.loaded = True
            print(f"\n✅ Dataset loaded successfully!")
            print(f"   Total claims: {len(self.claims):,}")
            print(f"   Verified claims (label=1): {sum(1 for c in self.claims if c['label'] == 1):,}")
            print(f"   Debunked claims (label=0): {sum(1 for c in self.claims if c['label'] == 0):,}")
            return True
            
        except Exception as e:
            logger.error(f"Error loading dataset: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def find_similar(self, query_text: str, threshold: float = 0.5, limit: int = 5) -> List[Dict]:
        """
        Find similar claims using string similarity
        Returns list of similar claims with their labels (0 or 1)
        """
        if not self.loaded:
            return []
        
        from difflib import SequenceMatcher
        
        query_lower = query_text.lower()
        
        # Calculate similarity for each claim
        similarities = []
        for claim in self.claims:
            claim_lower = claim['text'].lower()
            ratio = SequenceMatcher(None, query_lower, claim_lower).ratio()
            
            if ratio >= threshold:
                similarities.append((ratio, claim))
        
        # Sort by similarity (descending) and return top matches
        similarities.sort(reverse=True, key=lambda x: x[0])
        return [claim for _, claim in similarities[:limit]]


# Global dataset instance
_dataset = None


def get_dataset() -> DatasetLoader:
    """Get or create global dataset instance"""
    global _dataset
    if _dataset is None:
        _dataset = DatasetLoader()
    return _dataset


def load_dataset() -> bool:
    """Load the global dataset"""
    dataset = get_dataset()
    return dataset.load()
